"""
Excel writer for writing Application IDs and status back to the workbook.

Performance: keeps the workbook open in memory for the entire batch run,
eliminating the ~10 s load_workbook() overhead that used to happen on
every single row.  Disk saves are deferred and only happen every
SAVE_INTERVAL writes (or on flush/close), cutting total I/O dramatically.
"""

import time
import openpyxl
from utils.constants import Sheet1Col

MAX_RETRIES = 3
RETRY_DELAY = 3          # seconds between retries on PermissionError
SAVE_INTERVAL = 5        # flush to disk every N writes


class ExcelWriteSession:
    """
    Persistent write session that keeps the workbook in memory.

    Usage:
        writer = ExcelWriteSession("data.xlsx")
        writer.write_app_id(2, "APP-123")
        writer.write_app_id(3, "APP-456")
        writer.close()          # flushes remaining writes to disk
    """

    def __init__(self, file_path: str, save_interval: int = SAVE_INTERVAL):
        self.file_path = file_path
        self.save_interval = save_interval
        self._wb = openpyxl.load_workbook(file_path)
        self._ws = self._wb["Sheet1"]
        self._pending = 0        # writes since last disk save

    # ── public API ──────────────────────────────────────────────

    def write_app_id(self, row: int, app_id: str):
        """Write Application ID to column K and auto-flush if interval reached."""
        self._ws.cell(row=row, column=Sheet1Col.APPLICATION_ID, value=app_id)
        self._tick()

    def write_status(self, row: int, status: str):
        """Write status / error message to column K."""
        self._ws.cell(row=row, column=Sheet1Col.APPLICATION_ID, value=status)
        self._tick()

    def flush(self):
        """Force-save any pending writes to disk now."""
        if self._pending > 0:
            self._save()

    def close(self):
        """Flush pending writes and release the workbook."""
        try:
            self.flush()
        finally:
            try:
                self._wb.close()
            except Exception:
                pass

    # ── internals ───────────────────────────────────────────────

    def _tick(self):
        self._pending += 1
        if self._pending >= self.save_interval:
            self._save()

    def _save(self):
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                self._wb.save(self.file_path)
                self._pending = 0
                return
            except PermissionError:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                else:
                    raise PermissionError(
                        f"Cannot write to '{self.file_path}' — file is locked. "
                        f"Please CLOSE the Excel file and try again."
                    )
