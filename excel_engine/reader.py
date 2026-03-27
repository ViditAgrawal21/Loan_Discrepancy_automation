"""
Excel reader for loading discrepancy data from Sheet1.
Uses openpyxl for precise cell-level access with column index mapping.
Single sheet only — no Survey_Data sheet needed for discrepancy management.
"""

import openpyxl
from utils.constants import Sheet1Col


def load_workbook(file_path: str):
    """Load the Excel workbook and return the Sheet1 worksheet."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except PermissionError:
        raise PermissionError(
            f"Cannot open '{file_path}' — file is locked. "
            f"Please CLOSE the Excel file and try again."
        )

    if "Sheet1" not in wb.sheetnames:
        raise ValueError("Excel file must contain a 'Sheet1' sheet")

    return wb["Sheet1"]


def get_total_rows(sheet1_ws) -> int:
    """Get the total number of data rows in Sheet1 (excluding header)."""
    max_row = sheet1_ws.max_row
    # Find actual last row with data in the Account column
    for row in range(max_row, 1, -1):
        if sheet1_ws.cell(row=row, column=Sheet1Col.ACCOUNT).value is not None:
            return row
    return 1  # Only header


def _cell_str(ws, row, col) -> str:
    """Get cell value as a stripped string, or empty string if None."""
    val = ws.cell(row=row, column=col).value
    if val is None:
        return ""
    return str(val).strip()


def _cell_val(ws, row, col):
    """Get raw cell value."""
    return ws.cell(row=row, column=col).value


def read_row(ws, row: int) -> dict:
    """
    Read a single row from Sheet1 and return as a dictionary.
    Row is 1-based (row 2 = first data row).
    """
    return {
        "sr": _cell_val(ws, row, Sheet1Col.SR),
        "aadhaar": _cell_str(ws, row, Sheet1Col.AADHAAR),
        "dob": _cell_str(ws, row, Sheet1Col.DOB),
        "account": _cell_str(ws, row, Sheet1Col.ACCOUNT),
        "name": _cell_str(ws, row, Sheet1Col.NAME),
        "village": _cell_str(ws, row, Sheet1Col.VILLAGE),
        "loan_section_date": _cell_str(ws, row, Sheet1Col.LOAN_SECTION_DATE),
        "loan_sanctioned": _cell_str(ws, row, Sheet1Col.LOAN_SANCTIONED),
        "kcc_drawing_limit": _cell_str(ws, row, Sheet1Col.KCC_DRAWING_LIMIT),
        "loan_sanctioned_activity": _cell_str(ws, row, Sheet1Col.LOAN_SANCTIONED_ACTIVITY),
        "application_id": _cell_str(ws, row, Sheet1Col.APPLICATION_ID),

        # Personal details (columns M-X)
        "name_as_per_aadhaar": _cell_str(ws, row, Sheet1Col.NAME_AS_PER_AADHAAR),
        "aadhaar_no_full": _cell_str(ws, row, Sheet1Col.AADHAAR_NO_FULL),
        "name_as_per_passbook": _cell_str(ws, row, Sheet1Col.NAME_AS_PER_PASSBOOK),
        "dob_form": _cell_str(ws, row, Sheet1Col.DOB_FORM),
        "gender": _cell_str(ws, row, Sheet1Col.GENDER),
        "mobile": _cell_str(ws, row, Sheet1Col.MOBILE),
        "social_category": _cell_str(ws, row, Sheet1Col.SOCIAL_CATEGORY),
        "farmer_category": _cell_str(ws, row, Sheet1Col.FARMER_CATEGORY),
        "farmer_type": _cell_str(ws, row, Sheet1Col.FARMER_TYPE),
        "primary_occupation": _cell_str(ws, row, Sheet1Col.PRIMARY_OCCUPATION),
        "relative_type": _cell_str(ws, row, Sheet1Col.RELATIVE_TYPE),
        "relative_name": _cell_str(ws, row, Sheet1Col.RELATIVE_NAME),

        # Residential address (columns Y-AC)
        "state": _cell_str(ws, row, Sheet1Col.STATE),
        "district": _cell_str(ws, row, Sheet1Col.DISTRICT),
        "block_subdistrict": _cell_str(ws, row, Sheet1Col.BLOCK_SUBDISTRICT),
        "village_residential": _cell_str(ws, row, Sheet1Col.VILLAGE_RESIDENTIAL),
        "pincode": _cell_str(ws, row, Sheet1Col.PINCODE),

        # Season
        "season": _cell_str(ws, row, Sheet1Col.SEASON),
    }


def has_application_id(ws, row: int) -> bool:
    """Check if a row already has an Application ID (already processed)."""
    val = ws.cell(row=row, column=Sheet1Col.APPLICATION_ID).value
    return val is not None and str(val).strip() != ""
